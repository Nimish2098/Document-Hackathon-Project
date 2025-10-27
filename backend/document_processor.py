import pypdf
import os 
import google.generativeai as genai
import numpy as np
from dotenv import load_dotenv
import docx
import openpyxl
# load api key

load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")

if not API_KEY:
    raise   ValueError("Google Api Key not found")

genai.configure(api_key=API_KEY)


# In-memory "Vector database"
db = {}

def get_embeddings(text:str, task_type:str)->list[float]:
    try:
        result = genai.embed_content(
            model="models/text-embedding-004",
            content=text,
            task_type=task_type
        )
        return result['embedding']
    except Exception as e:
        print(f"Error in generating embeddings {e}")
        return []

def cosine_similarity(vec1,vec2):
    dot_product = np.dot(vec1,vec2)
    norm_vec1=np.linalg.norm(vec1)
    norm_vec2=np.linalg.norm(vec2)
    if norm_vec1==0 or norm_vec2==0:
        return 0.0
    return dot_product/(norm_vec1*norm_vec2)
# processing the documents
def process_document(file_path:str,file_name:str)->bool:
    print(f"Processing Document:{file_name} from {file_path}")
    text = ""
    try:
        file_ext = file_name.lower().split('.')[-1]
         
        
        if file_ext=="pdf":
            reader = pypdf.PdfReader(file_path)
            
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text
            print(f"Extracted {len(text)} characters from {file_name}")
            
        elif file_ext=="docx":
            doc = docx.Document(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
                
        elif file_ext=="xlsx":
            workbook  = openpyxl.load_workbook(file_path)
            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                for row in worksheet.iter_rows():
                    for cell in worksheet.iter_cols():
                        if cell.value:
                            text+=str(cell.value)+" "
                        text += "\n"
        else :
            print(f"File type: {file_name.split('.')[-1]} not supported yet")
            return False
        
        if not text:
            print("No Text Extracted, skipping.")
            return False
        
        chunks = text.split("\n\n")
        
        chunks = [chunk.strip() for chunk in chunks if chunk.strip()]
        print(f"Split text into {len(chunks)} chunks.")
        
        if not chunks:
            print("No text chunks found")
            return False
      
        print(f"Generate embeddings for {len(chunks)} chunls")
        processed_chunks=[]
        embeddings = []
        for i, chunk in enumerate(chunks):
            print(f"embedding chunk {i+1}/{len(chunks)}")
            emb = get_embeddings(chunk, task_type="RETRIEVAL_DOCUMENT")
            if emb:
                embeddings.append(emb)
                processed_chunks.append(chunk)
            else:
                print(f"Skipping chunk {i+1}")
        chunks = processed_chunks   
        print(f"Generated {len(embeddings)} real embeddings")
        
        if not embeddings:
            print("No embeddings could be generated.")
            return False
        
        if "documind_collection" not in db:
            db["documind_collection"] = {
                "chunks":[],
                "embeddings":[],
                "metadata":[]
            }
        collection = db["documind_collection"]  
        for i, chunk in enumerate(chunks):
            collection["chunks"].append(chunk)
            collection["embeddings"].append(embeddings[i])
            collection["metadata"].append({
                 "source_file":file_name,
                "chunk_id":i
            })

        print(f"Successfully indexed {file_name}")
        return True
    
    except Exception as e:
        print(f"Error processing documents {file_name} : {e}")
        return False
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)
    

def query_documents(query_text:str)-> (str,str):
    print(f"Received Query: {query_text}")
    
    try:
        if "documind_collection" not in db  or not db["documind_collection"]["chunks"]:
            return "No Document have been indexed"

        print("Generating query embeddings")
        query_embeddings = get_embeddings(query_text, task_type="RETRIEVAL_QUERY")
        if not query_embeddings:
            return "Could not generate embeddings for query"
        
        
        collection = db["documind_collection"]
        all_doc_embeddings = collection["embeddings"]
        
        similarities = [cosine_similarity(query_embeddings,emb) for emb in all_doc_embeddings]
        
        top_k = 3
        top_k_indices = np.argsort(similarities)[-top_k:][::-1]
        relevant_chunks = [collection["chunks"][i] for i in top_k_indices]
        relevant_metadata = [collection["metadata"][i] for i in top_k_indices]
        
        if not relevant_chunks:
            return "Could not find any relevant information"
        
        print(f"Found {len(relevant_chunks)} relevant chunks")
        
        context  = "\n\n---\n\n".join(relevant_chunks)
        
        prompt = f"""
        CONTEXT:
        {context}
        Question:
        {query_text}
        """
        
        print("Built prompt for LLM.")
        
        print("Sending prompt to gemini")
        model = genai.GenerativeModel('gemini-2.5-flash-preview-09-2025')
        response = model.generate_content(prompt)
                    
        answer = response.text
        
        source_file = relevant_metadata[0]["source_file"] if relevant_metadata else "unknown"

        return answer, source_file
    except Exception as e:
        print(f"UNEXPECTED ERROR {e}")
        
        return f"An internal error occurred: {str(e)}", "Error"
            
    